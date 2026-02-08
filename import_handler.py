"""
ITAM - Import Handler
엑셀 Import/Export 처리 모듈
"""
import sqlite3
import json
from datetime import datetime
from io import BytesIO

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("⚠️ openpyxl 미설치. pip install openpyxl 실행 필요")

DB_PATH = 'itam.db'

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn

# ========================================
# 검증 함수
# ========================================
def validate_required(value, field_name):
    if not value or str(value).strip() == '':
        return f"필수 필드 '{field_name}' 누락"
    return None

def validate_date(value, field_name):
    if not value:
        return None
    try:
        if isinstance(value, datetime):
            return None
        datetime.strptime(str(value), '%Y-%m-%d')
        return None
    except:
        return f"'{field_name}' 날짜 형식 오류 (YYYY-MM-DD)"

def validate_number(value, field_name):
    if not value:
        return None
    try:
        float(value)
        return None
    except:
        return f"'{field_name}' 숫자 형식 오류"

def validate_location(conn, location_name):
    row = conn.execute('SELECT location_id FROM Location WHERE location_name = ? AND is_active = 1', 
                       (location_name,)).fetchone()
    return row['location_id'] if row else None

def validate_category(conn, category_name):
    row = conn.execute('SELECT category_id FROM AssetCategory WHERE category_name = ? AND is_active = 1', 
                       (category_name,)).fetchone()
    return row['category_id'] if row else None

def validate_user_by_empno(conn, employee_no):
    row = conn.execute('SELECT user_id FROM User WHERE employee_no = ?', (employee_no,)).fetchone()
    return row['user_id'] if row else None

def validate_asset_status(status):
    valid_statuses = ['신규', '사용중', '여유', '수리중', '폐기예정', '폐기', '분실']
    return status in valid_statuses

# ========================================
# HW 자산 Import
# ========================================
def import_hw_assets(file_content):
    """HW 자산 Import 처리"""
    conn = get_db()
    wb = load_workbook(filename=BytesIO(file_content))
    ws = wb.active
    
    results = {'success': 0, 'errors': [], 'warnings': []}
    headers = [cell.value for cell in ws[1]]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # 빈 행 스킵
            continue
            
        data = dict(zip(headers, row))
        errors = []
        warnings = []
        
        # 필수 필드 검증
        if err := validate_required(data.get('asset_name'), '자산명'):
            errors.append(err)
        if err := validate_required(data.get('category_name'), '자산 카테고리'):
            errors.append(err)
        if err := validate_required(data.get('asset_status'), '상태'):
            errors.append(err)
        if err := validate_required(data.get('location_name'), '사업장'):
            errors.append(err)
        if err := validate_required(data.get('manager_employee_no'), '관리담당자사번'):
            errors.append(err)
        
        # 날짜 검증
        if err := validate_date(data.get('purchase_date'), '구매일'):
            errors.append(err)
        if err := validate_date(data.get('warranty_start'), '보증시작일'):
            errors.append(err)
        if err := validate_date(data.get('warranty_end'), '보증만료일'):
            errors.append(err)
        
        # 참조 데이터 검증
        location_id = validate_location(conn, data.get('location_name'))
        if not location_id and data.get('location_name'):
            errors.append(f"사업장 '{data.get('location_name')}' 시스템에 없음")
        
        category_id = validate_category(conn, data.get('category_name'))
        if not category_id and data.get('category_name'):
            errors.append(f"카테고리 '{data.get('category_name')}' 시스템에 없음")
        
        manager_id = validate_user_by_empno(conn, data.get('manager_employee_no'))
        if not manager_id and data.get('manager_employee_no'):
            errors.append(f"관리담당자 사번 '{data.get('manager_employee_no')}' 시스템에 없음")
        
        user_id = None
        if data.get('employee_no'):
            user_id = validate_user_by_empno(conn, data.get('employee_no'))
            if not user_id:
                warnings.append(f"사용자 사번 '{data.get('employee_no')}' 시스템에 없음 (배정 보류)")
        
        if not validate_asset_status(data.get('asset_status', '')):
            errors.append(f"상태 '{data.get('asset_status')}' 허용 값 아님")
        
        # 자산번호 중복 체크
        if data.get('asset_number'):
            existing = conn.execute('SELECT 1 FROM Asset WHERE asset_number = ?', 
                                    (data.get('asset_number'),)).fetchone()
            if existing:
                errors.append(f"자산번호 '{data.get('asset_number')}' 중복")
        
        if errors:
            results['errors'].append({'row': row_idx, 'asset_name': data.get('asset_name'), 'errors': errors})
            continue
        
        # 자산번호 자동 채번
        asset_number = data.get('asset_number')
        if not asset_number:
            loc = conn.execute('SELECT location_code FROM Location WHERE location_id = ?', (location_id,)).fetchone()
            cat = conn.execute('SELECT category_code FROM AssetCategory WHERE category_id = ?', (category_id,)).fetchone()
            year = datetime.now().year
            prefix = f"{loc['location_code']}-{cat['category_code']}-{year}"
            last = conn.execute("SELECT asset_number FROM Asset WHERE asset_number LIKE ? ORDER BY asset_number DESC LIMIT 1", (f"{prefix}-%",)).fetchone()
            if last:
                num = int(last['asset_number'].split('-')[-1]) + 1
            else:
                num = 1
            asset_number = f"{prefix}-{str(num).zfill(4)}"
        
        # 스펙 정보 조합
        specs = {}
        if data.get('spec_cpu'): specs['cpu'] = data['spec_cpu']
        if data.get('spec_ram_gb'): specs['ram_gb'] = data['spec_ram_gb']
        if data.get('spec_storage'): specs['storage'] = data['spec_storage']
        
        # 날짜 포맷
        def fmt_date(v):
            if isinstance(v, datetime):
                return v.strftime('%Y-%m-%d')
            return str(v) if v else None
        
        # Insert
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO Asset (asset_number, asset_name, category_id, asset_status, location_id, install_location,
                              manufacturer, model_name, serial_number, specifications, purchase_date, purchase_cost,
                              warranty_start, warranty_end, useful_life_months, current_user_id,
                              asset_manager_id, ip_address, mac_address, hostname, os_info, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        ''', (asset_number, data['asset_name'], category_id, data['asset_status'], location_id,
              data.get('install_location'), data.get('manufacturer'), data.get('model_name'),
              data.get('serial_number'), json.dumps(specs, ensure_ascii=False) if specs else None,
              fmt_date(data.get('purchase_date')), data.get('purchase_cost'),
              fmt_date(data.get('warranty_start')), fmt_date(data.get('warranty_end')),
              data.get('useful_life_months'), user_id, manager_id,
              data.get('ip_address'), data.get('mac_address'), data.get('hostname'),
              data.get('os_info'), data.get('notes')))
        
        results['success'] += 1
        if warnings:
            results['warnings'].append({'row': row_idx, 'asset_name': data.get('asset_name'), 'warnings': warnings})
    
    conn.commit()
    conn.close()
    return results

# ========================================
# SW 라이선스 Import
# ========================================
def import_licenses(file_content):
    """SW 라이선스 Import 처리"""
    conn = get_db()
    wb = load_workbook(filename=BytesIO(file_content))
    ws = wb.active
    
    results = {'success': 0, 'errors': [], 'warnings': []}
    headers = [cell.value for cell in ws[1]]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
            
        data = dict(zip(headers, row))
        errors = []
        
        if err := validate_required(data.get('software_name'), '소프트웨어명'):
            errors.append(err)
        if err := validate_required(data.get('category_name'), '카테고리'):
            errors.append(err)
        if err := validate_required(data.get('license_type'), '라이선스유형'):
            errors.append(err)
        if err := validate_required(data.get('license_metric'), '측정단위'):
            errors.append(err)
        if err := validate_required(data.get('total_quantity'), '총수량'):
            errors.append(err)
        if err := validate_required(data.get('manager_employee_no'), '관리담당자사번'):
            errors.append(err)
        
        category_id = validate_category(conn, data.get('category_name'))
        if not category_id and data.get('category_name'):
            errors.append(f"카테고리 '{data.get('category_name')}' 시스템에 없음")
        
        manager_id = validate_user_by_empno(conn, data.get('manager_employee_no'))
        if not manager_id and data.get('manager_employee_no'):
            errors.append(f"관리담당자 사번 '{data.get('manager_employee_no')}' 시스템에 없음")
        
        if errors:
            results['errors'].append({'row': row_idx, 'software_name': data.get('software_name'), 'errors': errors})
            continue
        
        license_number = data.get('license_number')
        if not license_number:
            year = datetime.now().year
            prefix = f"HQ-SW-{year}"
            last = conn.execute("SELECT license_number FROM SoftwareLicense WHERE license_number LIKE ? ORDER BY license_number DESC LIMIT 1", (f"{prefix}-%",)).fetchone()
            num = int(last['license_number'].split('-')[-1]) + 1 if last else 1
            license_number = f"{prefix}-{str(num).zfill(4)}"
        
        def fmt_date(v):
            if isinstance(v, datetime):
                return v.strftime('%Y-%m-%d')
            return str(v) if v else None
        
        total = int(data['total_quantity'])
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO SoftwareLicense (license_number, software_name, category_id, version, license_type,
                                         license_metric, total_quantity, used_quantity, available_quantity,
                                         purchase_date, purchase_cost, subscription_start, subscription_end,
                                         is_subscription, auto_renewal, license_manager_id, compliance_status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (license_number, data['software_name'], category_id, data.get('version'),
              data['license_type'], data['license_metric'], total, total,
              fmt_date(data.get('purchase_date')), data.get('purchase_cost'),
              fmt_date(data.get('subscription_start')), fmt_date(data.get('subscription_end')),
              1 if data.get('subscription_end') else 0, 1 if str(data.get('auto_renewal')).upper() == 'Y' else 0,
              manager_id, '정상', data.get('notes')))
        
        results['success'] += 1
    
    conn.commit()
    conn.close()
    return results

# ========================================
# Export 함수
# ========================================
def export_assets():
    """자산 목록 Export"""
    conn = get_db()
    rows = conn.execute('''
        SELECT a.asset_number, a.asset_name, c.category_name, a.asset_status, l.location_name,
               a.install_location, a.manufacturer, a.model_name, a.serial_number, a.specifications,
               a.purchase_date, a.purchase_cost, a.warranty_start, a.warranty_end, a.useful_life_months,
               u.employee_no, u.user_name, m.employee_no as manager_employee_no, m.user_name as manager_name,
               a.ip_address, a.mac_address, a.hostname, a.os_info, a.notes
        FROM Asset a
        LEFT JOIN AssetCategory c ON a.category_id = c.category_id
        LEFT JOIN Location l ON a.location_id = l.location_id
        LEFT JOIN User u ON a.current_user_id = u.user_id
        LEFT JOIN User m ON a.asset_manager_id = m.user_id
        WHERE a.is_deleted = 0
        ORDER BY a.asset_id
    ''').fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "자산목록"
    
    headers = ['자산번호', '자산명', '카테고리', '상태', '사업장', '상세위치', '제조사', '모델명', '시리얼번호',
               '스펙', '구매일', '구매금액', '보증시작일', '보증만료일', '사용연한(월)', '사용자사번', '사용자이름',
               '관리담당자사번', '관리담당자', 'IP주소', 'MAC주소', '호스트명', '운영체제', '비고']
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(dict(row).values(), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_licenses():
    """라이선스 목록 Export"""
    conn = get_db()
    rows = conn.execute('''
        SELECT sl.license_number, sl.software_name, c.category_name, sl.version, sl.license_type,
               sl.license_metric, sl.total_quantity, sl.used_quantity, sl.available_quantity,
               sl.purchase_date, sl.purchase_cost, sl.subscription_start, sl.subscription_end,
               sl.auto_renewal, m.employee_no as manager_employee_no, m.user_name as manager_name,
               sl.compliance_status, sl.notes
        FROM SoftwareLicense sl
        LEFT JOIN AssetCategory c ON sl.category_id = c.category_id
        LEFT JOIN User m ON sl.license_manager_id = m.user_id
        WHERE sl.is_deleted = 0
    ''').fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "라이선스목록"
    
    headers = ['관리번호', '소프트웨어명', '카테고리', '버전', '유형', '측정단위', '총수량', '사용', '가용',
               '구매일', '구매금액', '구독시작', '구독만료', '자동갱신', '담당자사번', '담당자', '상태', '비고']
    
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(dict(row).values(), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ========================================
# 템플릿 생성
# ========================================
def create_hw_template():
    """HW 자산 Import 템플릿 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "HW자산템플릿"
    
    headers = ['asset_number', 'asset_name', 'category_name', 'asset_status', 'location_name',
               'install_location', 'manufacturer', 'model_name', 'serial_number',
               'spec_cpu', 'spec_ram_gb', 'spec_storage', 'purchase_date', 'purchase_cost',
               'warranty_start', 'warranty_end', 'useful_life_months',
               'employee_no', 'user_name', 'manager_employee_no',
               'ip_address', 'mac_address', 'hostname', 'os_info', 'notes']
    
    headers_kr = ['자산번호(선택)', '자산명*', '카테고리*', '상태*', '사업장*',
                  '상세위치', '제조사', '모델명', '시리얼번호',
                  'CPU', 'RAM(GB)', '저장장치', '구매일', '구매금액',
                  '보증시작일', '보증만료일', '사용연한(월)',
                  '사용자사번', '사용자이름', '관리담당자사번*',
                  'IP주소', 'MAC주소', '호스트명', '운영체제', '비고']
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, (en, kr) in enumerate(zip(headers, headers_kr), 1):
        cell = ws.cell(row=1, column=col, value=en)
        cell.fill = header_fill
        cell.font = header_font
        ws.cell(row=2, column=col, value=kr)
    
    # 예시 데이터
    example = ['', '홍길동 노트북', '노트북', '사용중', '본사', 'A동 3층', 'Lenovo', 'ThinkPad T14s',
               'PF12345', 'i7-13700', '16', 'SSD 512GB', '2024-03-15', '1500000',
               '2024-03-15', '2027-03-14', '36', 'EMP001', '홍길동', 'EMP100',
               '192.168.1.100', 'AA:BB:CC:DD:EE:FF', 'PC-HONG-01', 'Windows 11', '']
    for col, val in enumerate(example, 1):
        ws.cell(row=3, column=col, value=val)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_license_template():
    """SW 라이선스 Import 템플릿 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "라이선스템플릿"
    
    headers = ['license_number', 'software_name', 'category_name', 'version', 'license_type',
               'license_metric', 'total_quantity', 'purchase_date', 'purchase_cost',
               'subscription_start', 'subscription_end', 'auto_renewal', 'manager_employee_no', 'notes']
    
    headers_kr = ['관리번호(선택)', '소프트웨어명*', '카테고리*', '버전', '유형*',
                  '측정단위*', '총수량*', '구매일', '구매금액',
                  '구독시작일', '구독만료일', '자동갱신(Y/N)', '관리담당자사번*', '비고']
    
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, (en, kr) in enumerate(zip(headers, headers_kr), 1):
        cell = ws.cell(row=1, column=col, value=en)
        cell.fill = header_fill
        cell.font = header_font
        ws.cell(row=2, column=col, value=kr)
    
    example = ['', 'Microsoft 365 E3', '오피스/생산성', '2024', '구독',
               'per_user', '100', '2024-01-01', '50000000',
               '2024-01-01', '2025-12-31', 'Y', 'EMP100', '']
    for col, val in enumerate(example, 1):
        ws.cell(row=3, column=col, value=val)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
